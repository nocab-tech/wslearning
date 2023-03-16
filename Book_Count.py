import openpyxl

# Open the workbook
workbook = openpyxl.load_workbook('book_info.xlsx')

# Select the sheet you want to read from
sheet = workbook.active

# Define a dictionary to store the categories and their counts
categories = {}

# Iterate through each row in the sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Extract the category from the row
    category = row[0]
    
    # Add the category to the dictionary if it's not already there
    if category not in categories:
        categories[category] = 0
    
    # Increment the count for the category
    categories[category] += 1

# Define a variable to keep track of the total count of books
total_books = 0

# Iterate through each row in the sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Extract the category from the row
    category = row[0]
    
    # Add the category to the dictionary if it's not already there
    if category not in categories:
        categories[category] = 0
    
    # Increment the count for the category and the total count
    categories[category] += 1
    total_books += 1

# Sort the categories by count, in descending order
sorted_categories = sorted(categories.items(), key=lambda x: x[1], reverse=True)

# Print the sorted categories and their counts, with a divider between each category
for category, count in sorted_categories:
    print(f"{category}: {count} books")
    print("--------------------")

# Print the total count of books
print(f"Total books: {total_books}")